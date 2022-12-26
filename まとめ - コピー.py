
import requests
from bs4 import BeautifulSoup
import clipboard
import openpyxl
import os


wb = openpyxl.load_workbook("****.xlsx")
ws = wb.worksheets[0]

url_list = []

for cell in ws['A']:
    url_list.append(cell.value)

#print(url_list)


# Webページを取得して解析する
for cell in ws['1']:
    #print(url_list[0])
    #load_url = url_list[0]
    load_url = cell.value
    html = requests.get(load_url)
    soup = BeautifulSoup(html.content, "html.parser")

    # title、Classを検索して、その文字列を表示する
    title = soup.find("title").text
    title2 = title.replace("****", '')
    print('<h2 class="p-magazine-article-header__title">'+title2+'</h2>')

    img = soup.find(class_="p-magazine-article-description__photo")

    elem = img.find("img")
    src = elem.get("src")
    print('<p><img src="'+src+'" alt="">')


    caption = soup.find(class_="p-magazine-article-description__caption")
    if not caption == None:
        print(caption)
    
    print('</p>')


    chap = soup.find(class_="p-magazine-article-description__text")
    print(chap)
    
    print('<p style="text-align: center;"><a href="'+load_url+'" class="c-btn-common c-btn-common--wrap c-btn-common">続きを読む</a></p>')
    print("<p></p>")
    print("<hr />")
    

os.remove('****.xlsx')





